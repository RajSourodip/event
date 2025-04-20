from flask import Flask, make_response, request, jsonify, send_from_directory, render_template, session, redirect, url_for, send_file, Response
from flask_bcrypt import Bcrypt
from openpyxl import Workbook
import openpyxl
import yagmail
from bson.json_util import dumps, ObjectId
from datetime import date, datetime, timedelta,  timezone
from pymongo import MongoClient
from dotenv import load_dotenv
from flask_cors import  CORS
from flask import Flask, render_template, request, redirect, url_for, session
import io
import pandas as pd

import os
load_dotenv()



myPASS = os.getenv('pass')
app = Flask(__name__)
app.secret_key = "hihihihih"

cors =  CORS(app)

app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=1)
# MongoDB connection
# app.config["MONGO_URI"] = "mongodb+srv://sourodip:rajghosh@first.ff1ia.mongodb.net/first?retryWrites=true&w=majority"
client = MongoClient("mongodb+srv://sourodip:rajghosh@first.ff1ia.mongodb.net/first?retryWrites=true&w=majority")
bcrypt = Bcrypt()
# Schemas: MongoDB collections
User = client.first.login
Record = client.first.records
Dash = client.first.dash
Event = client.first.events
Admin = client.first.admin
Permit  = client.first.permissions

@app.route('/SubmitDistrictReport', methods=['POST', 'GET'])
def submit_district_report():
    # Get the JSON data from the request
    data = request.get_json()

    if "username" in session:
        data["user"] = session["username"]
    else:
        return jsonify({"status": "error", "message": "User not logged in"}), 401

    # Add date and time
    data["date"] = datetime.now().strftime("%Y-%m-%d")
    data["time"] = datetime.now().strftime("%H:%M:%S")

    try:
        # Insert the data into the Record collection
        Dash.insert_one(data)
        print("successfully inserted: ")
        print(data)

        # Send a success response
        return jsonify({"status": "success", "message": "Report submitted successfully"}), 200
    except Exception as e:
        # Handle exceptions and send an error response
        print(f"An error occurred: {e}")
        return jsonify({"status": "error", "message": "Failed to submit report"}), 500

@app.route('/download')
def download_data():
    # Fetch data from MongoDB
    data = list(Record.find({}, {'_id': False}))

    # Convert data to DataFrame
    df = pd.DataFrame(data)

    # Convert DataFrame to Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)

    # Return the Excel file
    return send_file(output, download_name="data.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')




@app.route('/downloads')
def download_records():

    records = list(Record.find({}, {'_id': False}))  # Exclude the _id field
    
    df = pd.DataFrame(records)
    
    # Create a BytesIO object and save the DataFrame to it as an Excel file
    output =  io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    # Rewind the BytesIO object
    output.seek(0)
    
    # Send the file
    return send_file(output, download_name="data.xlsx", as_attachment=True,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/report_form', methods=['GET'])
def report_form():
    district = request.args.get('district')  # Get the district from the query string
    
    if not district:
        return jsonify({"status": "error", "message": "District not specified"}), 400
  
    return render_template("report_form.html", district=district)


# @app.route('/getReportsByDistrict', methods=['GET'])
# def get_reports_by_district():
#     # Ensure the user is authenticated by checking session data
#     username = session.get('username')  # Assuming the user_id is stored in the session
#     print(f"User ID from session: {username}")  # Debugging: Check if the user is authenticated
    
#     if not username:
        
#         return redirect(url_for('login')), 401

#     # Fetch all reports for the current user
#     query = {"user": username}  # Only fetch reports for the authenticated user
#     print(f"Query to fetch reports: {query}")  # Debugging: Print the query to see if it's correct

#     try:
#         reports_data = list(Record.find(query))  # Fetch reports from the database
#         print(reports_data)
#     except Exception as e:
#         print(f"Error while fetching reports: {str(e)}")  # Log any database fetching errors
#         return jsonify({"status": "error", "message": "Error fetching reports from the database"}), 500

#     if not reports_data:
#         print("No reports found for the user.")  # Debugging: If no reports are found
#         return jsonify({"status": "error", "message": "No reports found for this user"}), 404

#     # Classify reports by district
#     reports_by_district = {}
    
#     for report in reports_data:
#         district = report.get('district', 'Unknown')  # Get district from report, default to 'Unknown' if not present
#         if district not in reports_by_district:
#             reports_by_district[district] = []
#         reports_by_district[district].append(report)

#     # Convert ObjectId to string for JSON serialization
#     for district, reports in reports_by_district.items():
#         for report in reports:
#             report["_id"] = str(report["_id"])

#     # Return the classified reports
#     print("Reports classified by district:", reports_by_district)  # Debugging: Check the classified reports
#     return jsonify({"status": "success", "data": reports_by_district}), 200

@app.route('/individual_records')
def individual_records():
    if 'username' in session:
        username = session['username']
        
        return render_template('individual.html')
    else:
        return redirect(url_for('index'))

@app.route('/get_permissions', methods=['GET'])
def get_permissions():
    try:
        print("Trying to get in")
        # Fetch all permission requests from the 'permissions' collection
        permissions_cursor = Permit.find({'status': 'pending'})
        print(permissions_cursor)
        # Create a list to hold the permission data
        permissions_list = []
#        
        
        # Iterate over each permission request and format the response
        for permission in permissions_cursor:
            permissions_list.append({
                'user': permission['user'],
                'request_date': permission['request_date'],
                'status': permission['status']
            })
        
        print("Printing:", permissions_list)  # Correct way to print the list
        # Return the list of permission requests as a JSON response
        return jsonify(permissions_list)
    
    except Exception as e:
        print(e)
        return jsonify({'message': 'Error fetching permissions', 'error': str(e)}), 500


# @app.route('/send_permission_request', methods=['POST', 'GET'])
# def send_permission_request():
#     try:
#         # Get the username from the incoming JSON request body
#         data = request.json
#         username = session.get("username")
#         # Check if username is provided
#         if not username:
#             return jsonify({'message': 'Username is required'}), 400

#         # Check if the user already has a request in the permissions collection
#         existing_request = Permit.find_one({'user': username})

#         # If the user has a granted request, return an error
#         if existing_request and existing_request.get('status') == 'granted':
#             return jsonify({'message': 'Permission already granted for this user'}), 400

#         # Create a new permission request document in the MongoDB database
#         if not existing_request or existing_request.get('status') == 'rejected':
#             now_utc = datetime.now(timezone.utc)
#             # Convert UTC to IST (UTC + 5:30)
#             ist_offset = timedelta(hours=5, minutes=30)
#             now_ist = now_utc + ist_offset
#             permission_request = {
#                 'user': username,
#                 'status': 'pending',  # Pending status initially
#                 'request_date': now_ist.isoformat()  # Use timezone-aware datetime
#             }
#             print(permission_request)
#             Permit.insert_one(permission_request)
#             return jsonify({'message': 'Permission request sent successfully', 'success': True}), 200

#         # Return success response
#         return jsonify({'message': 'Permission request already exists for this user'}), 400

#     except Exception as e:
#         print(f"Error: {str(e)}")
#         return jsonify({'message': 'Error sending permission request', 'error': str(e)}), 500

@app.route('/send_permission_request', methods=['POST', 'GET'])
def send_permission_request():
    try:
        # Get current user from session
        username = session.get("username")
        if not username:
            return jsonify({'message': 'Username is required'}), 400

        # Check if user already has a request
        existing_request = Permit.find_one({'user': username})

        if existing_request:
            if existing_request.get('status') == 'granted':
                return jsonify({'message': 'Permission already granted for this user'}), 400
            if existing_request.get('status') == 'pending':
                return jsonify({'message': 'Permission request already exists for this user'}), 400

        # Create permission request
        now_utc = datetime.now(timezone.utc)
        ist_offset = timedelta(hours=5, minutes=30)
        now_ist = now_utc + ist_offset

        permission_request = {
            'user': username,
            'status': 'pending',
            'request_date': now_ist.isoformat()
        }

        Permit.insert_one(permission_request)
        return jsonify({'message': 'Permission request sent successfully', 'success': True}), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'message': 'Error sending permission request', 'error': str(e)}), 500
    
@app.route('/check_permission', methods=['POST', 'GET'])
def check_permission():
    try:
        print("Received POST request for /check_permission")  # Debug print
        username = session.get("username")
        if not username:
            return jsonify({'message': 'Username is required'}), 400
        
        permission = Permit.find_one({'user': username})

        if permission:
            if permission['status'] == 'granted':
                return jsonify({'hasAccess': True}), 200
            else:
                return jsonify({'hasAccess': False}), 200
        else:
            return jsonify({'hasAccess': False}), 200

    except Exception as e:
        print(f"Error: {str(e)}")  # Debug error message
        return jsonify({'message': 'Error checking permission', 'error': str(e)}), 500

@app.route('/update_permission', methods=['POST'])
def update_permission():
    try:
        # Get data from the POST request
        username = request.json.get('username')
        status = request.json.get('status')  # Expected values: 'granted' or 'rejected'
        
        # Validate the input
        if status not in ['granted', 'rejected']:
            return jsonify({'message': 'Invalid status. Use "granted" or "rejected".'}), 400
        
        # Update the permission status in the 'permissions' collection
        result = Permit.update_one(
            {'user': username, 'status': 'pending'},  # Find pending request by username
            {'$set': {'status': status}}  # Set the new status
        )
        
        if result.modified_count > 0:
            return jsonify({'message': f'Permission {status} successfully!'}), 200
        else:
            return jsonify({'message': 'Permission request not found or already updated.'}), 404
    
    except Exception as e:
        # Handle unexpected errors
        return jsonify({'message': 'Error updating permission', 'error': str(e)}), 500


@app.route('/signup', methods=['POST'])
def signup():
    # Get the data from the POST request
    username = request.json.get('username')
    password = request.json.get('password')

    # Check if the username already exists in the 'admin' collection
    if Admin.find_one({'username': username}):
        return jsonify({'message': 'Username already exists'}), 400

    # Hash the password before storing it
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')

    # Store the new admin's credentials in the 'admin' collection
    Admin.insert_one({
        'username': username,
        'password': hashed_password
    })

    return jsonify({'message': 'Signup successful!'}), 201

@app.route('/signin', methods=['POST'])
def signin():
    # Get the data from the POST request
    username = request.json.get('username')
    password = request.json.get('password')

    # Find the admin by username in the 'admin' collection
    admin = Admin.find_one({'username': username})

    if admin and bcrypt.check_password_hash(admin['password'], password):
        return jsonify({'message': 'Login successful!'}), 200
    else:
        return jsonify({'message': 'Invalid username or password'}), 401

@app.route('/admin')
def admin():
    return render_template("admin.html")

# @app.route('/getWeeklyReport', methods=['GET'])
# def get_weekly_report():
#     # Get the current date and time
#     now = datetime.now()
#     today = datetime.now().date()
#     start_of_month = today.replace(day=1)
#     totals = {}
#     # Calculate the date for one week ago
#     one_week_ago = now - timedelta(days=7)
#     username = session.get("username")
#     if "username" in session:
#         username = session["username"]
#     # Query MongoDB to find records created within the last week
#         # user_reports = Record.find({"user": username, "demoGivenDate": {"$gte": one_week_ago.strftime('%Y-%m-%d')}})
#         user_reports = Record.find({"user": username})

#         # Prepare the response data
#         reports_data = []
#         for report in user_reports:
#             # Convert ObjectId to string for compatibility
#             report["_id"] = str(report["_id"])
            
#             # Convert datetime fields to string for JSON serialization
#             for key in ["demoGivenDate", "followUpDate", "signedAt"]:
#                 if key in report:
#                     report[key] = report[key].strftime("%Y-%m-%d %H:%M:%S")
            
#             # Calculate PD Done Today for this report
#             pd_done_today = Record.count_documents({
#                 "PD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(today, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             print("pd done today ", pd_done_today)

#             smd_done_today = Record.count_documents({
#                 "SMD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(today, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             print("smd done today ", smd_done_today)
                      
#             MTD_smd_done = Record.count_documents({
#                 "SMD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
            
#             print("mtd smd done ", pd_done_today)  

#             MTD_approached = Record.count_documents({
#                 "remarksStatus": "Approached",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             MTD_pd_appointment = Record.count_documents({
#                 "remarksStatus": "PD appointments",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })

#             MTD_pd_feedback_collected = Record.count_documents({
#                 "remarksStatus": "PD feedback form collected",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             MTD_signup_done = Record.count_documents({
#                 "remarksStatus": "Sign up done",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })

#             MTD_pd_done = Record.count_documents({
#                 "PD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             totals["pdDoneToday"] = pd_done_today
#             totals["smdDoneToday"] = smd_done_today
#             totals["MTDpdDone"] = MTD_pd_done
#             totals["MTDsmdDone"] = MTD_smd_done

#             totals["mtdApproached"] = MTD_approached
#             totals["mtdPdAppointments"] = MTD_pd_appointment
#             totals["mtdPdFeedbackFormCollected"] = MTD_pd_feedback_collected
#             totals["mtdSignupDone"] = MTD_signup_done

#             # Add the calculated fields to the report
#             report["pdDoneToday"] = pd_done_today
#             report["smdDoneToday"] = smd_done_today
#             report["MTDpdDone"] = MTD_pd_done
#             report["MTDsmdDone"] = MTD_smd_done
#             report["mtdApproached"] = MTD_approached
#             report["mtdPdAppointments"] = MTD_pd_appointment
#             report["mtdPdFeedbackFormCollected"] = MTD_pd_feedback_collected
#             report["mtdSignupDone"] = MTD_signup_done
#             print("The final reports are: ",  str(report))
            
#             # Append the enriched report to the list
#             reports_data.append(report)
#         print(reports_data)
#         # Return the enriched reports
#         return jsonify({"status": "success", "data": reports_data, "totals":totals}), 200

#     else:
#         return jsonify({"status": "error", "message": "User not logged in"}), 401


@app.route('/getWeeklyReport', methods=['GET'])
def get_weekly_report():
    now = datetime.now()
    today = datetime.now().date()
    start_of_month = today.replace(day=1)
    totals = {}
    one_week_ago = now - timedelta(days=7)
    username = session.get("username")

    if username:
        user_reports = Record.find({"user": username})
        reports_data = []
        
        
        totals = {}  # To store aggregated counts for the report
        for report in user_reports:
            report["_id"] = str(report["_id"])
            for key in ["demoGivenDate", "followUpDate", "signedAt"]:
                if key in report:
                    report[key] = report[key].strftime("%Y-%m-%d %H:%M:%S")
            totals = {}
            # Aggregated counts
            totals["pdDoneToday"] = Record.count_documents({
                "PD": True,
                "demoGivenDate": {
                    "$gte": datetime.combine(today, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            totals["smdDoneToday"] = Record.count_documents({
                "SMD": True,
                "demoGivenDate": {
                    "$gte": datetime.combine(today, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            # MTD (Month-To-Date) counts
            totals["mtdSmdDone"] = Record.count_documents({
                "SMD": True,
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            totals["mtdApproached"] = Record.count_documents({
                "remarksStatus": "Approached",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            totals["mtdPdAppointments"] = Record.count_documents({
                "remarksStatus": "PD appointments",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            totals["mtdPdFeedbackFormCollected"] = Record.count_documents({
                "remarksStatus": "PD feedback form collected",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            totals["mtdSignupDone"] = Record.count_documents({
                "remarksStatus": "Sign up done",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            totals["mtdPdDone"] = Record.count_documents({
                "PD": True,
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            # Append the report with totals data if needed
            # report["totals"] = totals
            # reports_data.append(report)

            report.update(totals)
            reports_data.append(report)

        return jsonify({"status": "success", "data": reports_data, "totals": totals}), 200
    else:
        return jsonify({"status": "error", "message": "User not logged in"}), 401


@app.route('/downloadWeeklyReport', methods=['GET'])
def download_weekly_report():
    now = datetime.now()
    today = datetime.now().date()
    start_of_month = today.replace(day=1)
    one_week_ago = now - timedelta(days=7)
    username = session.get("username")

    if "username" in session:
        username = session["username"]

        # Query MongoDB for the reports
        user_reports = Record.find({"user": username})

        # Create a new Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Report"

        # Define headers
        headers = [
            "Report ID", "Date", "Schools Approached", "Week No.", 
            "PD Done Today", "SMD Done Today", "Signed", "Signed At",
            "MTD Approached", "MTD PD Appointments", "MTD PD Done",
            "MTD SMD Done", "MTD Signup Done", "Submitted by"
        ]
        ws.append(headers)

        # Populate Excel rows with report data
        for report in user_reports:
            # Convert ObjectId to string and handle date fields
            report_id = str(report.get("_id", ""))
            demo_given_date = report.get("demoGivenDate", "").strftime("%Y-%m-%d") if report.get("demoGivenDate") else ""
            schools_approached = "Approached" if report.get("remarksStatus") == "Approached" else "Not Approached"
            week_no = report.get("weekSelect", "")
            pd_done = "PD Done" if report.get("PD") else "PD Not Done"
            smd_done = "SMD Done" if report.get("SMD") else "SMD Not Done"
            signed = "Signed" if report.get("remarksStatus") == "Signed" else "Not Signed"
            signed_at = report.get("signedAt", "").strftime("%Y-%m-%d %H:%M:%S") if report.get("signedAt") else ""

            mtd_approached = report.get("mtdApproached", 0)
            mtd_pd_appointments = report.get("mtdPdAppointments", 0)
            mtd_pd_done = report.get("MTDpdDone", 0)
            mtd_smd_done = report.get("MTDsmdDone", 0)
            mtd_signup_done = report.get("mtdSignupDone", 0)
            submitted_by = report.get("user", "")

            # Add a row to the worksheet
            ws.append([
                report_id, demo_given_date, schools_approached, week_no,
                pd_done, smd_done, signed, signed_at,
                mtd_approached, mtd_pd_appointments, mtd_pd_done,
                mtd_smd_done, mtd_signup_done, submitted_by
            ])

        # Save the Excel workbook to a BytesIO stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create a response to download the file
        filename = f"Weekly_Report_{username}_{today}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    else:
        return jsonify({"status": "error", "message": "User not logged in"}), 401
    



    
# @app.route('/downloadWeeklyReport', methods=['GET'])
# def download_weekly_report():
#     username = session.get("username")

#     if username:
#         user_reports = Record.find({"user": username})

#         # Create an Excel workbook and worksheet
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.title = "Weekly Report"

#         # Add headers to the worksheet
#         headers = [
#             "ID", "Demo Given Date", "Follow Up Date", "Signed At",
#             "Remarks Status", "PD Done", "SMD Done", "User"
#         ]
#         sheet.append(headers)

#         # Add data rows to the worksheet
#         for report in user_reports:
#             row = [
#                 report.get("_id"),
#                 report.get("demoGivenDate", ""),
#                 report.get("followUpDate", ""),
#                 report.get("signedAt", ""),
#                 report.get("remarksStatus", ""),
#                 "Yes" if report.get("PD") else "No",
#                 "Yes" if report.get("SMD") else "No",
#                 report.get("user", "")
#             ]
#             sheet.append(row)

#         # Save the workbook to a BytesIO stream
#         output = io.BytesIO()
#         workbook.save(output)
#         output.seek(0)

#         # Create a response with the Excel file
#         response = make_response(output.read())
#         response.headers["Content-Disposition"] = "attachment; filename=weekly_report.xlsx"
#         response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         return response

#     else:
#         return jsonify({"status": "error", "message": "User not logged in"}), 401
    

# @app.route('/getReport', methods=['GET'])
# def get_report():
#     if "username" in session:
#         username = session["username"]

#         # Get today's date and the start of the current month
#         today = date.today()
#         start_of_month = date(today.year, today.month, 1) 

#         # Query the database for all reports by the logged-in user
#         user_reports = Record.find({"user": username})
        
#         # Prepare the reports data
#         reports_data = []
#         for report in user_reports:
#             # Convert ObjectId to string for compatibility
#             report["_id"] = str(report["_id"])
            
#             # Convert datetime fields to string for JSON serialization
#             for key in ["demoGivenDate", "followUpDate", "signedAt"]:
#                 if key in report:
#                     report[key] = report[key].strftime("%Y-%m-%d %H:%M:%S")
            
#             # Calculate PD Done Today for this report
#             pd_done_today = Record.count_documents({
#                 "PD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(today, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             smd_done_today = Record.count_documents({
#                 "SMD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(today, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
            
#             # Calculate SMD Done Today for this report
#             MTD_smd_done = Record.count_documents({
#                 "SMD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
            
#             MTD_approached = Record.count_documents({
#                 "remarksStatus": "Approached",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             MTD_pd_appointment = Record.count_documents({
#                 "remarksStatus": "PD appointments",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })

#             MTD_pd_feedback_collected = Record.count_documents({
#                 "remarksStatus": "PD feedback form collected",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             MTD_signup_done = Record.count_documents({
#                 "remarksStatus": "Sign up done",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })

#             MTD_pd_done = Record.count_documents({
#                 "PD": True,  # Filter where PD is True
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),  # Start of month
#                     "$lte": datetime.combine(today, datetime.max.time())           # End of today
#                 },
#                 "user": username  # Specific user filter
#             })
            
#             # Add the calculated fields to the report
#             report["pdDoneToday"] = pd_done_today
#             report["smdDoneToday"] = smd_done_today
#             report["MTDpdDone"] = MTD_pd_done
#             report["MTDsmdDone"] = MTD_smd_done
#             report["mtdApproached"] = MTD_approached
#             report["mtdPdAppointments"] = MTD_pd_appointment
#             report["mtdPdFeedbackFormCollected"] = MTD_pd_feedback_collected
#             report["mtdSignupDone"] = MTD_signup_done
#             print("The final reports are: ",  str(report))
            
#             # Append the enriched report to the list
#             reports_data.append(report)
#             print(reports_data)
#         # Return the enriched reports
#         return jsonify({"status": "success", "data": reports_data}), 200

#     else:
#         return jsonify({"status": "error", "message": "User not logged in"}), 401

@app.route('/getReport', methods=['GET'])
def get_report():
    if "username" not in session:
        return jsonify({"status": "error", "message": "User not logged in"}), 401
    
    username = session["username"]
    today = date.today()
    start_of_month = date(today.year, today.month, 1)
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())
    start_of_month_start = datetime.combine(start_of_month, datetime.min.time())

    try:
        # Fetch all reports for the user
        user_reports = list(Record.find({"user": username}))
        if not user_reports:
            return jsonify({"status": "error", "message": "No reports found"}), 404

        reports_data = []
        for report in user_reports:
            report["_id"] = str(report["_id"])
            for key in ["demoGivenDate", "followUpDate", "signedAt"]:
                if key in report:
                    report[key] = report[key].strftime("%Y-%m-%d %H:%M:%S")

            # Calculate metrics
            report["pdDoneToday"] = count_metric("PD", True, today_start, today_end, username)
            report["smdDoneToday"] = count_metric("SMD", True, today_start, today_end, username)
            report["MTDpdDone"] = count_metric("PD", True, start_of_month_start, today_end, username)
            report["MTDsmdDone"] = count_metric("SMD", True, start_of_month_start, today_end, username)
            report["mtdApproached"] = count_metric("remarksStatus", "Approached", start_of_month_start, today_end, username)
            report["mtdPdAppointments"] = count_metric("remarksStatus", "PD appointments", start_of_month_start, today_end, username)
            report["mtdPdFeedbackFormCollected"] = count_metric("remarksStatus", "PD feedback form collected", start_of_month_start, today_end, username)
            report["mtdSignupDone"] = count_metric("remarksStatus", "Sign up done", start_of_month_start, today_end, username)

            reports_data.append(report)

        return jsonify({"status": "success", "data": reports_data}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

def count_metric(field, value, start_date, end_date, username):
    return Record.count_documents({
        field: value,
        "demoGivenDate": {"$gte": start_date, "$lte": end_date},
        "user": username
    })





# @app.route('/downloadIndiReport', methods=['GET'])
# def download_Indi_report():
#     if "username" in session:
#         username = session["username"]
#         user_reports = Record.find({"user": username})
#         today = datetime.now().date()
#         start_of_month = today.replace(day=1)

#         # Convert MongoDB Cursor to List of Dictionaries
#         reports_data = []
#         for report in user_reports:
#             MTD_smd_done = Record.count_documents({
#                 "SMD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
            
#             MTD_approached = Record.count_documents({
#                 "remarksStatus": "Approached",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             MTD_pd_appointment = Record.count_documents({
#                 "remarksStatus": "PD appointments",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })

#             MTD_pd_feedback_collected = Record.count_documents({
#                 "remarksStatus": "PD feedback form collected",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
#             MTD_signup_done = Record.count_documents({
#                 "remarksStatus": "Sign up done",
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })

#             MTD_pd_done = Record.count_documents({
#                 "PD": True,
#                 "demoGivenDate": {
#                     "$gte": datetime.combine(start_of_month, datetime.min.time()),
#                     "$lte": datetime.combine(today, datetime.max.time())
#                 },
#                 "user": username  # Specific to this user
#             })
            
#             surveyDone = "Yes" if report.get("surveyDone") == True else "No"
#             report_dict = {
#                 "Report ID": str(report["_id"]),
#                 "Today's Report": report.get("demoGivenDate", ""),
#                 "Submitted by": report.get("user", ""),
#                 "Date": report.get("demoGivenDate", ""),
#                 "Survey Done": surveyDone,
#                 "Appointment Fixed": report.get("appointmentFixed", ""),
#                 "PD Done Today": "Yes" if report.get("PD") == True else "No",
#                 "SMD Done Today": "Yes" if report.get("SMD") == True else "No",
#                 "Signed": "Yes" if report.get("remarksStatus") == "Signed" else "No",
#                 "Signed At": report.get("signedAt", ""),
#                 "MTD Approached": MTD_approached,
#                 "MTD PD Appointments": MTD_pd_appointment,
#                 "MTD PD Done": MTD_pd_done,
#                 "MTD PD Feedback Form Collected": MTD_pd_feedback_collected,
#                 "MTD SMD Done": MTD_smd_done,
#                 "MTD Signup Done": MTD_signup_done,
#                 "Week No": report.get("weekSelect", ""),
#             }
#             reports_data.append(report_dict)

#         # Convert data to Pandas DataFrame
#         df = pd.DataFrame(reports_data)

#         # Convert DataFrame to Excel
#         excel_buffer = pd.ExcelWriter("individual_records.xlsx", engine="openpyxl")
#         df.to_excel(excel_buffer, index=False, sheet_name="Reports")
#         excel_buffer.close()

#         # Send file to frontend
#         with open("individual_records.xlsx", "rb") as f:
#             file_data = f.read()

#         response = Response(file_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         response.headers["Content-Disposition"] = "attachment; filename=individual_records.xlsx"
#         return response

#     else:
#         return jsonify({"status": "error", "message": "User not logged in"}), 401
    

@app.route('/downloadIndiReport', methods=['GET'])
def download_Indi_report():
    if "username" in session:
        username = session["username"]
        user_reports = Record.find({"user": username})
        today = datetime.now().date()
        start_of_month = today.replace(day=1)

        reports_data = []
        for report in user_reports:
            MTD_smd_done = Record.count_documents({
                "SMD": True,
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            MTD_approached = Record.count_documents({
                "remarksStatus": "Approached",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            MTD_pd_appointment = Record.count_documents({
                "remarksStatus": "PD appointments",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            MTD_pd_feedback_collected = Record.count_documents({
                "remarksStatus": "PD feedback form collected",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            MTD_signup_done = Record.count_documents({
                "remarksStatus": "Sign up done",
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            MTD_pd_done = Record.count_documents({
                "PD": True,
                "demoGivenDate": {
                    "$gte": datetime.combine(start_of_month, datetime.min.time()),
                    "$lte": datetime.combine(today, datetime.max.time())
                },
                "user": username
            })

            surveyDone = "Yes" if report.get("surveyDone") == True else "No"
            report_dict = {
                "Report ID": str(report["_id"]),
                "Today's Report": report.get("demoGivenDate", ""),
                "Submitted by": report.get("user", ""),
                "Date": report.get("demoGivenDate", ""),
                "Survey Done": surveyDone,
                "Appointment Fixed": report.get("appointmentFixed", ""),
                "PD Done Today": "Yes" if report.get("PD") == True else "No",
                "SMD Done Today": "Yes" if report.get("SMD") == True else "No",
                "Signed": "Yes" if report.get("remarksStatus") == "Signed" else "No",
                "Signed At": report.get("signedAt", ""),
                "MTD Approached": MTD_approached,
                "MTD PD Appointments": MTD_pd_appointment,
                "MTD PD Done": MTD_pd_done,
                "MTD PD Feedback Form Collected": MTD_pd_feedback_collected,
                "MTD SMD Done": MTD_smd_done,
                "MTD Signup Done": MTD_signup_done,
                "Week No": report.get("weekSelect", ""),
            }
            reports_data.append(report_dict)

    #     # Convert to DataFrame
    #     df = pd.DataFrame(reports_data)

    #     # Create in-memory Excel file
    #     output = BytesIO()
    #     with pd.ExcelWriter(output, engine='openpyxl') as writer:
    #         df.to_excel(writer, index=False, sheet_name='Reports')

    #     output.seek(0)  # Go to beginning of the BytesIO buffer

    #     # Return the file as a download response
    #     return Response(
    #         output,
    #         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #         headers={"Content-Disposition": "attachment;filename=individual_records.xlsx"}
    #     )

    # else:
    #     return jsonify({"status": "error", "message": "User not logged in"}), 401

    # Convert data to Pandas DataFrame
        df = pd.DataFrame(reports_data)

        # Convert DataFrame to Excel
        excel_buffer = pd.ExcelWriter("individual_records.xlsx", engine="openpyxl")
        df.to_excel(excel_buffer, index=False, sheet_name="Reports")
        excel_buffer.close()

        # Send file to frontend
        with open("individual_records.xlsx", "rb") as f:
            file_data = f.read()

        response = Response(file_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = "attachment; filename=individual_records.xlsx"
        return response

    else:
        return jsonify({"status": "error", "message": "User not logged in"}), 401
    
@app.route('/getReportsByDistrict', methods=['GET'])
def get_reports_by_district():
    username = session.get('username')
    print(f"User ID from session: {username}")

    if not username:
        return redirect(url_for('login')), 401

    query = {"user": username}
    print(f"Query to fetch reports: {query}")

    try:
        reports_data = list(Record.find(query))
        print(reports_data)
    except Exception as e:
        print(f"Error while fetching reports: {str(e)}")
        return jsonify({"status": "error", "message": "Error fetching reports from the database"}), 500

    if not reports_data:
        print("No reports found for the user.")
        return jsonify({"status": "error", "message": "No reports found for this user"}), 404

    reports_by_district = {}
    for report in reports_data:
        district = report.get('district', 'Unknown')
        if district not in reports_by_district:
            reports_by_district[district] = []
        reports_by_district[district].append(report)

    for district, reports in reports_by_district.items():
        for report in reports:
            report["_id"] = str(report["_id"])

    print("Reports classified by district:", reports_by_district)
    return jsonify({"status": "success", "data": reports_by_district}), 200

@app.route('/downloadReportsByDistrict', methods=['GET'])
def download_reports_by_district():
    username = session.get('username')
    print(f"User ID from session: {username}")

    if not username:
        return redirect(url_for('login')), 401

    query = {"user": username}
    print(f"Query to fetch reports: {query}")

    try:
        reports_data = list(Record.find(query))
        print(reports_data)
    except Exception as e:
        print(f"Error while fetching reports: {str(e)}")
        return jsonify({"status": "error", "message": "Error fetching reports from the database"}), 500

    if not reports_data:
        print("No reports found for the user.")
        return jsonify({"status": "error", "message": "No reports found for this user"}), 404

    reports_by_district = {}
    for report in reports_data:
        district = report.get('district', 'Unknown')
        if district not in reports_by_district:
            reports_by_district[district] = []
        reports_by_district[district].append(report)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "District Reports"

    # Adding headers
    headers = ["District", "Schools", "PD", "SMD", "Signed", "SMD to Sign Up Follow Up", "Hot Follow Up", "NI"]
    sheet.append(headers)

    for district, reports in reports_by_district.items():
        stats = {
            "schools": len(reports),
            "pd": sum(1 for r in reports if r.get('PD')),
            "smd": sum(1 for r in reports if r.get('SMD')),
            "signed": sum(1 for r in reports if r.get('remarksStatus') == "Signed"),
            "smdToSignUpFollowUp": sum(1 for r in reports if r.get('remarksStatus') == "SMD Follow UP"),
            "hotFollowUp": sum(1 for r in reports if r.get('remarksStatus') == "Hot Follow Up"),
            "ni": sum(1 for r in reports if r.get('remarksStatus') == "NI"),
        }

        row = [
            district,
            stats["schools"],
            stats["pd"],
            stats["smd"],
            stats["signed"],
            stats["smdToSignUpFollowUp"],
            stats["hotFollowUp"],
            stats["ni"],
        ]
        sheet.append(row)

    # Save the workbook to a BytesIO stream
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=district_reports.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response    
    



@app.route('/')
def index():
    return render_template("index.html")
@app.route('/show')
def show():
    return render_template("show.html")
@app.route('/record')
def record():
    return render_template("record.html")
@app.route('/weekly_records')
def wr():
    return render_template('weekly.html')
@app.route('/calendar')
def calendar():
    return render_template("calendar.html")
@app.route('/dashboard')
def dashboard():
    return render_template("dashboard.html")
@app.route('/district_records')
def dr():
    return render_template("district.html")

# Route to handle form submission

@app.route('/submitForm', methods=['POST'])
def submit_form():
    form_data = request.json
    print(form_data)
    # Convert string dates to datetime objects
    form_data['demoGivenDate'] = datetime.strptime(form_data['demoGivenDate'], '%Y-%m-%d')
    form_data['followUpDate'] = datetime.strptime(form_data['followUpDate'], '%Y-%m-%d')
    form_data['signedAt'] = datetime.now()

    # Insert the record into the collection
    result = Record.insert_one(form_data)
    print(form_data)
    return jsonify({'message': 'Form submitted successfully', 'id': str(result.inserted_id)}), 201


# Endpoint to fetch report data

@app.route('/fetch', methods=['GET'])
def fetch_user_records():
    try:
        # Get the username from the query parameters or headers (depending on your approach)
        username = session.get('username')  # Assuming the username is passed as a query parameter
        
        if not username:
            return jsonify({'message': 'Username is required'}), 400

        # Query the database for records for the specific user
        user_records = list(Record.find({'user': username}))
        
        # If no records are found, return a message
        if not user_records:
            return jsonify({'message': 'No records found for this user'}), 404

        # Format the records as a list of dictionaries
        records_list = []
        for record in user_records:
            # Optionally, remove the MongoDB-specific `_id` field
            record['_id'] = str(record['_id'])  # Convert ObjectId to string if needed
            records_list.append(record)

        return jsonify(records_list), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'message': 'Error fetching user records', 'error': str(e)}), 500


    
# Get events by date
@app.route('/fevents', methods=['GET'])
def get_events():
    day = request.args.get('day')
    month = request.args.get('month')
    year = request.args.get('year')
    username = session['username']
    events = list(Event.find({"user":username}))
    
    return dumps(events), 200


def convert_objectid_to_str(document):
    if isinstance(document, dict):
        return {k: convert_objectid_to_str(v) for k, v in document.items()}
    elif isinstance(document, list):
        return [convert_objectid_to_str(i) for i in document]
    elif isinstance(document, ObjectId):
        return str(document)
    else:
        return document

# @app.route('/events', methods=['POST', 'GET'])
# def add_event():
#     try:
#         event_data = request.get_json()
#         username = session.get("username")
#         event_data["user"] = username
#         print(event_data)
        
#         # Insert the event into MongoDB
#         Event.insert_one(event_data)
        
#         # Extract necessary fields from the event data
#         day = event_data.get("day")
#         month = event_data.get("month")
#         year = event_data.get("year")
#         school_name = event_data.get("schoolName")
#         district = event_data.get("district")
#         mobile_number = event_data.get("mobileNumber")
#         email = event_data.get("email")
#         admin_email = event_data.get("admin_email") 


#         message = (f"Hello,\n\nYour event at '{school_name}' has been successfully registered!\n\n"
#                    f"Event Details:\n"
#                    f"Date: {day}/{month}/{year}\n"
#                    f"School Name: {school_name}\n"
#                    f"District: {district}\n"
#                    f"Mobile Number: {mobile_number}\n\n"
#                    "Thank you for using our service!")
        
#         # Customize subject and content of the email
#         subject = "Event Registration Confirmation"
#         print(message)
        
#         # Send email using yagmail
#         yag = yagmail.SMTP('ghoshraj368@gmail.com', "wumwyxwyqfabhetz")
#         yag.send(
#             to=email,
#             subject=subject,
#             contents=message
#         )
        
#         # Return the event data with a 201 response
#         return jsonify({"status": "success"}), 201
#     except Exception as e:
#         print(e)
#         return jsonify({'error': str(e)}), 500

@app.route('/events', methods=['POST', 'GET'])
def add_event():
    try:
        event_data = request.get_json()
        username = session.get("username")
        event_data["user"] = username
        print(event_data)
        
        # Insert the event into MongoDB
        Event.insert_one(event_data)
        
        # Extract necessary fields from the event data
        day = event_data.get("day")
        month = event_data.get("month")
        year = event_data.get("year")
        school_name = event_data.get("schoolName")
        district = event_data.get("district")
        mobile_number = event_data.get("mobileNumber")
        email = event_data.get("email")
        admin_email = event_data.get("admin_email")
        
        message = (f"Hello,\n\nYour event at '{school_name}' has been successfully registered!\n\n"
                   f"Event Details:\n"
                   f"Date: {day}/{month}/{year}\n"
                   f"School Name: {school_name}\n"
                   f"District: {district}\n"
                   f"Mobile Number: {mobile_number}\n\n"
                   "Thank you for using our service!")
        
        # Customize subject and content of the email
        subject = "Event Registration Confirmation"
        print(message)
        
        # Send email using yagmail
        yag = yagmail.SMTP('ghoshraj368@gmail.com', 'wumwyxwyqfabhetz')
        yag.send(
            to=[email, admin_email],
            subject=subject,
            contents=message
        )
        
        # Return the event data with a 201 response
        return jsonify({"status": "success"}), 201
    except Exception as e:
        print(e)
        return jsonify({'error': str(e)}), 500

@app.route('/events', methods=['DELETE'])
def delete_event():
    try:
        event_data = request.get_json()
        Event.delete_one({
            "day": event_data['day'],
            "month": event_data['month'],
            "year": event_data['year'],
            "title": event_data['title']
        })
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# User login or registration
@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data['username']
    password = data['password']
    session["username"]  = username
    
    print("session username set to: ", username)
    print(password)

    user = User.find_one({"user": username})
    
    if user:
        if user['password'] == password:
            session['username'] =  username
            return jsonify({'success': True, 'message': 'Login successful'}), 200
        else:
            return jsonify({'success': False, 'message': 'Incorrect password'}), 401
    else:
        User.insert_one({"user": username, "password": password})
        return jsonify({'success': True, 'message': 'User registered and logged in'}), 201



@app.route('/tracker')
def tracker():
    if 'username' in session:
        return render_template('tracker.html')
    return redirect(url_for('index'))

@app.route('/update_location', methods=['POST'])
def update_location():
    if 'username' in session:
        latitude = request.form['latitude']
        longitude = request.form['longitude']
        # Here you can save the coordinates to a database or perform other operations
        return f"Location updated to {latitude}, {longitude}"
    return "User not logged in", 403
# Submit a new record


@app.route('/submit', methods=['POST', 'GET'])
def submit_record():
    try:  
        # Print incoming JSON data for debugging
        record_data = request.get_json()
        print('Received record data:', record_data)
        
        # Ensure required fields are present in the data
        if not all(key in record_data for key in ['demoGivenDate', 'followUpDate']):
            raise ValueError("Missing required fields")

        # Parse dates
        record_data['demoGivenDate'] = datetime.strptime(record_data['demoGivenDate'], '%Y-%m-%d')
        record_data['followUpDate'] = datetime.strptime(record_data['followUpDate'], '%Y-%m-%d')
        
        # Add current datetime for 'signedAt' and username from session
        record_data['signedAt'] = datetime.now()
        record_data["user"] = session.get("username")

        # Insert record into MongoDB
        Record.insert_one(record_data)
        
        return jsonify({'message': 'Record submitted successfully!'}), 200
    except ValueError as ve:
        # Handle missing required fields or invalid data
        return jsonify({'message': 'Error submitting record.', 'error': str(ve)}), 400
    except Exception as e:
        # General error handling
        print('Error:', e)
        return jsonify({'message': 'Error submitting record.', 'error': str(e)}), 500




# Fetch records data
@app.route('/fetch_records', methods=['GET'])
def fetch_records():
    try:
        username = session.get("username")
        records = list(Record.find({'user': username}))
        return dumps(records), 200
    except Exception as e:
        return jsonify({'message': 'Error fetching data.', 'error': str(e)}), 500

# Serve HTML files for specific routes
@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory('public', filename)

